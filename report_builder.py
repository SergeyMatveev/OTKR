from __future__ import annotations

import time
from datetime import datetime
from pathlib import Path
from typing import List, Dict
from logging import Logger
from decimal import Decimal, InvalidOperation
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


_PKO_RAW = "ПРОФЕССИОНАЛЬНАЯКОЛЛЕКТОРСКАЯОРГАНИЗАЦИЯ"
_PKO_PATTERN = re.compile("(?iu)" + r"\s*".join(list(_PKO_RAW)))


def _is_nd(value: object) -> bool:
    if value is None:
        return True
    s = str(value).strip()
    return s == "" or s.upper() == "Н/Д"


def _normalize_whitespace(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()


def _normalize_org_forms(name: str) -> str:
    """
    Нормализует ОПФ в наименовании:
      - длинные формы "ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ" -> "ООО "
      - "ПРОФЕССИОНАЛЬНАЯ КОЛЛЕКТОРСКАЯ ОРГАНИЗАЦИЯ" (с разрывами) -> "ПКО "
    """
    if not name:
        return "Н/Д"

    s = _normalize_whitespace(name)

    # ООО
    s = re.sub(
        r"(?iu)\bОБЩЕСТВ[О0]\s+С\s+ОГРАНИЧЕННОЙ\s+ОТВЕТСТВЕННОСТ[ЬЪB6]Ю?\b\s*",
        "ООО ",
        s,
    )

    # ПКО (допускаем разрывы внутри слов)
    s = _PKO_PATTERN.sub("ПКО ", s)

    # Убираем лишние пробелы перед кавычками
    s = re.sub(r"\s+([\"»])", r"\1", s)
    s = re.sub(r"\s+(«)", r"\1", s)
    s = _normalize_whitespace(s)

    return s or "Н/Д"


def _normalize_creditor_name(name: object, *, normalize_opf: bool) -> str:
    if name is None:
        return "Н/Д"
    s = str(name).strip()
    if not s or _is_nd(s):
        return "Н/Д"
    s = _normalize_whitespace(s)
    if normalize_opf:
        s = _normalize_org_forms(s)
    return s or "Н/Д"


def _normalize_inn(value: object) -> str:
    if value is None:
        return "Н/Д"
    s = str(value).strip()
    if not s or _is_nd(s):
        return "Н/Д"
    m = re.search(r"(?<!\d)(\d{10})(?!\d)", s)
    if m:
        return m.group(1)
    return "Н/Д"


def _normalize_page(value: object) -> str:
    if value is None:
        return "Н/Д"
    if isinstance(value, float) and pd.isna(value):
        return "Н/Д"
    s = str(value).strip()
    if not s or s.upper() == "Н/Д":
        return "Н/Д"
    try:
        n = int(float(s.replace(" ", "").replace(",", ".")))
        return str(n) if n > 0 else "Н/Д"
    except Exception:
        return s or "Н/Д"


def _normalize_amount(value: object) -> str:
    """
    Приводит строку с числом к формату ХХХХХ,YY.
    Если число не найдено или некорректно — возвращает "Н/Д".
    """
    if value is None:
        return "Н/Д"
    s = str(value).strip()
    if not s or s.upper() == "Н/Д":
        return "Н/Д"

    m = re.search(r"\d[\d ]*(?:[.,]\d{1,2})?", s)
    if not m:
        return "Н/Д"
    num = m.group(0)
    num = num.replace(" ", "")
    num = num.replace(",", ".")

    try:
        dec = Decimal(num)
    except InvalidOperation:
        return "Н/Д"

    dec = dec.quantize(Decimal("0.01"))
    s_val = format(dec, "f").replace(".", ",")
    return s_val


def _format_sum_with_currency(value: object) -> str:
    """
    Для строки "Сумма и валюта:":
      - если есть число -> "<сумма_2_знака> RUB"
      - если "Н/Д" -> "Н/Д"
    """
    norm = _normalize_amount(value)
    if norm == "Н/Д":
        return "Н/Д"
    return f"{norm} RUB"


def _format_debt(value: object) -> str:
    """
    Для строки "Текущая задолженность:":
      - если есть число -> "<сумма_2_знака>"
      - если "Н/Д" -> "Н/Д"
    """
    return _normalize_amount(value)


def _group_thousands_spaces(amount_str: str) -> str:
    if not amount_str:
        return "Н/Д"
    s = str(amount_str).strip()
    if not s or s.upper() == "Н/Д":
        return "Н/Д"

    if "," in s:
        int_part, frac_part = s.split(",", 1)
        has_frac = True
    else:
        int_part, frac_part = s, ""
        has_frac = False

    sign = ""
    digits = int_part
    if digits.startswith("-"):
        sign = "-"
        digits = digits[1:]

    if len(digits) <= 3:
        grouped = sign + digits
    else:
        parts: List[str] = []
        i = len(digits)
        while i > 0:
            start = max(0, i - 3)
            parts.append(digits[start:i])
            i = start
        grouped = sign + " ".join(reversed(parts))

    if has_frac:
        return f"{grouped},{frac_part}"
    return grouped


def _drop_fraction(amount_str: str) -> str:
    if not amount_str:
        return "Н/Д"
    s = str(amount_str).strip()
    if not s or s.upper() == "Н/Д":
        return "Н/Д"
    if "," in s:
        return s.split(",", 1)[0]
    return s


def _format_sum_with_currency_for_telegram(value: object) -> str:
    norm = _normalize_amount(value)
    if norm == "Н/Д":
        return "Н/Д"
    return f"{_group_thousands_spaces(norm)} RUB"


def _format_debt_for_telegram(value: object) -> str:
    norm = _normalize_amount(value)
    if norm == "Н/Д":
        return "Н/Д"
    return _group_thousands_spaces(norm)


def _format_sum_with_currency_for_excel(value: object) -> str:
    norm = _normalize_amount(value)
    if norm == "Н/Д":
        return "Н/Д"
    integer = _drop_fraction(norm)
    return f"{_group_thousands_spaces(integer)} RUB"


def _format_sum_for_excel_number(value: object) -> object:
    norm = _normalize_amount(value)
    if norm == "Н/Д":
        return None
    try:
        dec = Decimal(norm.replace(",", ".")).quantize(Decimal("0.01"))
        return float(dec)
    except Exception:
        return None


def _format_debt_for_excel(value: object) -> str:
    norm = _normalize_amount(value)
    if norm == "Н/Д":
        return "Н/Д"
    return norm.replace(",", ".")


def _format_debt_for_excel_number(value: object) -> object:
    norm = _normalize_amount(value)
    if norm == "Н/Д":
        return None
    try:
        dec = Decimal(norm.replace(",", ".")).quantize(Decimal("0.01"))
        return float(dec)
    except Exception:
        return None


def _postprocess_client_xlsx(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active

    widths = {
        "Страница": 9,
        "Название": 30,
        "ИНН": 11,
        "Дата сделки": 15,
        "Сумма кредита": 15,
        "Текущая задолженность": 25,
        "УИд договора": 40,
    }

    sum_col_idx = None
    debt_col_idx = None

    for cell in ws[1]:
        col_name = cell.value
        if col_name in widths:
            letter = get_column_letter(cell.col_idx)
            ws.column_dimensions[letter].width = widths[col_name]
        if col_name == "Сумма кредита":
            sum_col_idx = cell.col_idx
        if col_name == "Текущая задолженность":
            debt_col_idx = cell.col_idx

    def _coerce_to_number(c) -> None:
        v = c.value
        if v is None:
            return
        if isinstance(v, str):
            s = v.strip()
            if not s or s.upper() == "Н/Д":
                return
            s = s.replace(" ", "").replace(",", ".")
            try:
                dec = Decimal(s)
            except Exception:
                return
            dec = dec.quantize(Decimal("0.01"))
            c.value = float(dec)
        elif isinstance(v, (int, float, Decimal)):
            try:
                dec = Decimal(str(v)).quantize(Decimal("0.01"))
                c.value = float(dec)
            except Exception:
                pass
        c.number_format = "0.00"

    if sum_col_idx is not None:
        for row_idx in range(2, ws.max_row + 1):
            _coerce_to_number(ws.cell(row=row_idx, column=sum_col_idx))

    if debt_col_idx is not None:
        for row_idx in range(2, ws.max_row + 1):
            _coerce_to_number(ws.cell(row=row_idx, column=debt_col_idx))

    wb.save(xlsx_path)


def build_credit_report_from_csv(
    csv_path: Path,
    logger: Logger,
    *,
    telegram_user_id: str,
    telegram_username: str,
    request_id: str,
    file_name: str,
) -> tuple[List[str], Path]:
    """
    Загружает итоговый CSV, добавляет колонку "Решение",
    формирует текстовый отчёт по открытым кредитам и возвращает список сообщений + путь к xlsx.
    """
    csv_path = csv_path.resolve()
    overall_start = time.perf_counter()

    base_extra: Dict[str, object] = {
        "telegram_user_id": telegram_user_id,
        "telegram_username": telegram_username,
        "request_id": request_id,
        "model": "N/A",
        "api_key_id": "N/A",
        "file_name": file_name,
    }

    xlsx_path = csv_path.parent / f"{datetime.now():%Y-%m-%d-%H-%M-%S}-{Path(file_name).stem}.xlsx"
    excel_columns = [
        "Страница",
        "Название",
        "ИНН",
        "Дата сделки",
        "Сумма кредита",
        "Текущая задолженность",
        "УИд договора",
    ]

    logger.info(
        "Построение текстового отчёта по CSV %s начато.",
        csv_path,
        extra={**base_extra, "stage": "report_builder_start", "duration_seconds": 0},
    )

    try:
        try:
            df = pd.read_csv(csv_path, encoding="utf-8-sig")
        except Exception:
            df = pd.read_csv(csv_path, encoding="utf-8")
    except Exception:
        duration = time.perf_counter() - overall_start
        logger.error(
            "Не удалось загрузить CSV %s для построения отчёта.",
            csv_path,
            exc_info=True,
            extra={
                **base_extra,
                "stage": "report_builder_read_error",
                "duration_seconds": round(duration, 3),
            },
        )
        raise

    if "Решение" not in df.columns:
        df["Решение"] = ""
    else:
        df["Решение"] = df["Решение"].fillna("")

    decisions: Dict[object, str] = {}
    use_acquirer: Dict[object, bool] = {}
    for idx in df.index:
        decisions[idx] = ""
        use_acquirer[idx] = False

    closed_uids = set()
    for idx in df.index:
        row = df.loc[idx]
        prekr = str(row.get("Прекращение обязательства", "") or "").strip()
        uid_val = str(row.get("УИд договора", "") or "").strip()
        if (
            prekr in {"Надлежащее исполнение обязательства", "Прощение долга", "Иное основание"}
            and uid_val != ""
            and uid_val != "Не найдено"
        ):
            closed_uids.add(uid_val)

    # Шаг 1: базовый фильтр по "Прекращение обязательства"
    uid_groups: Dict[str, List[object]] = {}

    for idx in df.index:
        row = df.loc[idx]

        debt_val = row.get("Сумма задолженности", "")
        if _is_nd(debt_val):
            decisions[idx] = "Исключен. Задолженность Н/Д"
            use_acquirer[idx] = False
            continue

        prekr = str(row.get("Прекращение обязательства", "") or "").strip()
        uid_val = str(row.get("УИд договора", "") or "").strip()

        if (
            prekr == "Н/Д"
            and uid_val != ""
            and uid_val != "Не найдено"
            and uid_val in closed_uids
        ):
            decisions[idx] = "Исключен. Такой же Уид был закрыт"
            use_acquirer[idx] = False
            continue

        if prekr != "Н/Д":
            if prekr == "Надлежащее исполнение обязательства":
                decisions[idx] = "Исключен. Надлежащее исполнение обязательства"
            elif prekr == "Прощение долга":
                decisions[idx] = "Исключен. Прощение долга"
            elif prekr == "Иное основание":
                decisions[idx] = "Исключен. Иное основание"
            else:
                decisions[idx] = "Исключен. Не попал в правила отбора"

        if prekr == "Н/Д":
            if uid_val == "" or uid_val == "Не найдено":
                decisions[idx] = "Добавлен"
                use_acquirer[idx] = False
            else:
                uid_groups.setdefault(uid_val, []).append(idx)

    # Шаг 2: группировка по УИд договора
    for uid, idxs in uid_groups.items():
        if not idxs:
            continue

        if len(idxs) == 1:
            # Один договор с данным УИд
            idx = idxs[0]
            row = df.loc[idx]
            acq_name = row.get("Приобретатель прав кредитора", "")
            acq_inn = row.get("ИНН приобретателя прав кредитора", "")

            acq_name_is_nd = _is_nd(acq_name)
            acq_inn_norm = _normalize_inn(acq_inn)
            if not acq_name_is_nd and acq_inn_norm != "Н/Д":
                decisions[idx] = "Добавлен. Кредит выкуплен"
                use_acquirer[idx] = True
            else:
                decisions[idx] = "Добавлен"
                use_acquirer[idx] = False
        else:
            # Несколько строк с одинаковым УИд
            acq_names = {
                idx: df.loc[idx].get("Приобретатель прав кредитора", "") for idx in idxs
            }
            idxs_nd = [idx for idx in idxs if _is_nd(acq_names[idx])]

            if idxs_nd:
                # Вариант A: есть строка без приобретателя (Н/Д) — считаем исходным кредитором
                base_idx = idxs_nd[0]
                decisions[base_idx] = "Добавлен"
                use_acquirer[base_idx] = False

                for idx in idxs:
                    if idx == base_idx:
                        continue
                    if not decisions[idx]:
                        decisions[idx] = "Исключен. Одинаковый УИд, кредит выкуплен"
                    use_acquirer[idx] = False
            else:
                # Вариант Б: все строки с заполненным приобретателем — все добавляем
                for idx in idxs:
                    decisions[idx] = "Добавлен. Несколько строк с одинаковым УИд"
                    use_acquirer[idx] = True

    # Финальное заполнение "Решение" для всех строк
    for idx in df.index:
        if not decisions[idx]:
            prekr = str(df.loc[idx].get("Прекращение обязательства", "") or "").strip()
            if prekr == "Н/Д":
                decisions[idx] = "Исключен. Не попал в правила отбора"
            elif prekr == "Надлежащее исполнение обязательства":
                decisions[idx] = "Исключен. Надлежащее исполнение обязательства"
            elif prekr == "Прощение долга":
                decisions[idx] = "Исключен. Прощение долга"
            elif prekr == "Иное основание":
                decisions[idx] = "Исключен. Иное основание"
            else:
                decisions[idx] = "Исключен. Не попал в правила отбора"

    df["Решение"] = [decisions[idx] for idx in df.index]

    # Перезаписываем CSV тем же путём
    try:
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    except Exception:
        duration = time.perf_counter() - overall_start
        logger.error(
            "Не удалось сохранить обновлённый CSV с колонкой 'Решение': %s",
            csv_path,
            exc_info=True,
            extra={
                **base_extra,
                "stage": "report_builder_write_error",
                "duration_seconds": round(duration, 3),
            },
        )
        raise

    added_mask = df["Решение"].astype(str).str.startswith("Добавлен")
    added_count = int(added_mask.sum())
    total_rows = int(len(df))

    logger.info(
        "Колонка 'Решение' заполнена: всего строк=%d, добавлено=%d, исключено=%d.",
        total_rows,
        added_count,
        total_rows - added_count,
        extra={
            **base_extra,
            "stage": "report_builder_decisions_done",
            "duration_seconds": 0,
        },
    )

    if added_count == 0:
        duration = time.perf_counter() - overall_start
        logger.info(
            "Открытых кредитов по заданным правилам не найдено, текстовый отчёт пуст.",
            extra={
                **base_extra,
                "stage": "report_builder_done",
                "duration_seconds": round(duration, 3),
            },
        )
        df_excel = pd.DataFrame([], columns=excel_columns)
        df_excel.to_excel(xlsx_path, index=False)
        _postprocess_client_xlsx(xlsx_path)
        return [], xlsx_path

    # Группировка "добавленных" строк по финальному ИНН
    groups: Dict[str, Dict[str, object]] = {}
    order_final_inn: List[str] = []

    for idx in df.index:
        if not added_mask.loc[idx]:
            continue

        row = df.loc[idx]
        if use_acquirer.get(idx, False):
            raw_name = row.get("Приобретатель прав кредитора", "")
            raw_inn = row.get("ИНН приобретателя прав кредитора", "")
            final_name = _normalize_creditor_name(raw_name, normalize_opf=True)
            final_inn = _normalize_inn(raw_inn)
        else:
            raw_name = row.get("Короткое название", "")
            raw_inn = row.get("ИНН", "")
            final_name = _normalize_creditor_name(raw_name, normalize_opf=False)
            final_inn = _normalize_inn(raw_inn)

        if final_name == "Н/Д" and not _is_nd(raw_name):
            # Подчистим хотя бы пробелы
            final_name = _normalize_whitespace(raw_name)

        if not final_name:
            final_name = "Н/Д"
        if not final_inn:
            final_inn = "Н/Д"

        group_key = final_inn

        if group_key not in groups:
            groups[group_key] = {
                "final_name": final_name,
                "final_inn": final_inn,
                "rows": [],
            }
            order_final_inn.append(group_key)

        groups[group_key]["rows"].append(
            {
                "idx": idx,
                "number": row.get("Номер", ""),
                "uid": str(row.get("УИд договора", "") or ""),
                "page": _normalize_page(row.get("Страница", "Н/Д")),
                "date": str(row.get("Дата сделки", "") or "Н/Д"),
                "sum_text": row.get("Сумма и валюта", "Н/Д"),
                "debt_text": row.get("Сумма задолженности", "Н/Д"),
            }
        )

    # Формируем текстовые блоки
    from string import ascii_uppercase

    blocks: List[str] = []
    excel_rows: List[Dict[str, object]] = []
    block_index = 0
    total_debt = Decimal("0.00")
    total_sum = Decimal("0.00")

    for inn_key in order_final_inn:
        group = groups[inn_key]
        rows = group["rows"]

        # Сортировка внутри группы по "Номер" (как в исходном CSV)
        def _num_key(item: Dict[str, object]) -> int:
            num_raw = str(item.get("number", "") or "").strip()
            try:
                return int(num_raw.split()[0])
            except (ValueError, IndexError):
                return 10**9

        rows.sort(key=_num_key)

        block_index += 1
        final_name = group["final_name"]
        final_inn = group["final_inn"]

        for r in rows:
            page_val = r.get("page") or "Н/Д"
            date = r["date"] or "Н/Д"
            sum_val = _format_sum_for_excel_number(r["sum_text"])
            debt_val = _format_debt_for_excel_number(r["debt_text"])
            uid_val = r["uid"] or "Н/Д"

            debt_norm_for_total = _normalize_amount(r["debt_text"])
            if debt_norm_for_total != "Н/Д":
                try:
                    total_debt += Decimal(debt_norm_for_total.replace(",", "."))
                except Exception:
                    pass

            sum_norm_for_total = _normalize_amount(r["sum_text"])
            if sum_norm_for_total != "Н/Д":
                try:
                    total_sum += Decimal(sum_norm_for_total.replace(",", "."))
                except Exception:
                    pass

            excel_rows.append(
                {
                    "Страница": page_val,
                    "Название": final_name,
                    "ИНН": final_inn,
                    "Дата сделки": date,
                    "Сумма кредита": sum_val,
                    "Текущая задолженность": debt_val,
                    "УИд договора": uid_val,
                }
            )

        block_lines: List[str] = []
        block_lines.append(f"{block_index}. Кредитор:")
        block_lines.append(f"    Наименование: {final_name}")
        block_lines.append(f"    ИНН: {final_inn}")

        if len(rows) == 1:
            r = rows[0]
            page_val = r.get("page") or "Н/Д"
            date = r["date"] or "Н/Д"
            sum_str = _format_sum_with_currency_for_telegram(r["sum_text"])
            debt_str = _format_debt_for_telegram(r["debt_text"])
            uid_val = r["uid"] or "Н/Д"

            block_lines.append("    Договор:")
            block_lines.append(f"        Страница в ПДФ: {page_val}")
            block_lines.append(f"        Дата сделки: {date}")
            block_lines.append(f"        Сумма и валюта: {sum_str}")
            block_lines.append(f"        Текущая задолженность: {debt_str}")
            block_lines.append(f"        УИд договора: {uid_val}")
        else:
            for i, r in enumerate(rows):
                label = ascii_uppercase[i] if i < len(ascii_uppercase) else str(i + 1)
                page_val = r.get("page") or "Н/Д"
                date = r["date"] or "Н/Д"
                sum_str = _format_sum_with_currency_for_telegram(r["sum_text"])
                debt_str = _format_debt_for_telegram(r["debt_text"])
                uid_val = r["uid"] or "Н/Д"

                block_lines.append(f"    Договор {label}:")
                block_lines.append(f"        Страница в ПДФ: {page_val}")
                block_lines.append(f"        Дата сделки: {date}")
                block_lines.append(f"        Сумма и валюта: {sum_str}")
                block_lines.append(f"        Текущая задолженность: {debt_str}")
                block_lines.append(f"        УИд договора: {uid_val}")

            # Проверка на несколько строк с одинаковым УИд внутри блока
            uid_counts: Dict[str, int] = {}
            for r in rows:
                u = (r["uid"] or "").strip()
                if u and u != "Не найдено":
                    uid_counts[u] = uid_counts.get(u, 0) + 1
            if any(c > 1 for c in uid_counts.values()):
                block_lines.append("    (Внимание, несколько строк с одинаковым УИд)")

        block_text = "\n".join(block_lines).rstrip()
        blocks.append(block_text)

    total_debt_rounded = total_debt.quantize(Decimal("0.01"))
    total_sum_rounded = total_sum.quantize(Decimal("0.01"))
    excel_rows.append(
        {
            "Страница": "",
            "Название": "",
            "ИНН": "",
            "Дата сделки": "",
            "Сумма кредита": float(total_sum_rounded),
            "Текущая задолженность": float(total_debt_rounded),
            "УИд договора": "",
        }
    )

    # Разбиение блоков на сообщения по лимиту 4000 символов
    messages: List[str] = []
    current_blocks: List[str] = []

    for block in blocks:
        if not current_blocks:
            current_blocks = [block]
            continue

        candidate = "\n\n".join(current_blocks + [block])
        if len(candidate) <= 4000:
            current_blocks.append(block)
        else:
            messages.append("\n\n".join(current_blocks))
            current_blocks = [block]

    if current_blocks:
        messages.append("\n\n".join(current_blocks))

    total_debt_msg = _group_thousands_spaces(format(total_debt_rounded, "f").replace(".", ","))
    messages.append(
        f"Общая задолженность {total_debt_msg} рублей. Возможны ошибки, пожалуйста, проверяйте информацию."
    )

    duration = time.perf_counter() - overall_start
    logger.info(
        "Построение текстового отчёта завершено: блоков=%d, сообщений=%d.",
        len(blocks),
        len(messages),
        extra={
            **base_extra,
            "stage": "report_builder_done",
            "duration_seconds": round(duration, 3),
        },
    )

    df_excel = pd.DataFrame(excel_rows, columns=excel_columns)
    df_excel.to_excel(xlsx_path, index=False)
    _postprocess_client_xlsx(xlsx_path)

    return messages, xlsx_path
